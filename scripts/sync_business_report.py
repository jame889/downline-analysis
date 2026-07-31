#!/usr/bin/env python3
"""Download, validate, and sync the First Global SPS Business Report."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from import_data import process_file


BASE_URL = "https://www.myfirstglobal.com"
LOGIN_PAGE = f"{BASE_URL}/login"
LOGIN_URL = f"{BASE_URL}/login-process"
REPORT_PAGE = f"{BASE_URL}/my-office/business-report"
DOWNLOAD_URL = f"{BASE_URL}/my-office/api/business-report/excel"
BINARY_TREE_URL = f"{BASE_URL}/my-office/api/tree"
ROOT_MEMBER_ID = "900057"


def selected_months(explicit_month: str | None, include_previous: bool) -> list[str]:
    if explicit_month:
        return [explicit_month]
    now = datetime.now(ZoneInfo("Asia/Bangkok"))
    months = [now.strftime("%Y-%m")]
    if include_previous and now.day <= 3:
        months.append((now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m"))
    return months


def login(session: requests.Session, username: str, password: str) -> None:
    response = session.get(LOGIN_PAGE, timeout=30)
    response.raise_for_status()
    response = session.post(
        LOGIN_URL,
        data={"username": username, "password": password, "redirectUrl": ""},
        timeout=30,
    )
    response.raise_for_status()
    if "/my-office" not in response.url:
        raise RuntimeError(
            f"First Global login failed (final path: {response.url}, "
            f"content-type: {response.headers.get('content-type', 'unknown')})"
        )


def download(session: requests.Session, month: str) -> bytes:
    session.get(REPORT_PAGE, timeout=30).raise_for_status()
    response = session.get(
        DOWNLOAD_URL,
        params={"weekCnt": month.replace("-", ""), "type": "spon"},
        headers={"Accept": "application/octet-stream"},
        timeout=60,
    )
    response.raise_for_status()
    content = response.content
    if len(content) < 10_000 or not content.startswith(b"PK"):
        raise RuntimeError(f"Downloaded content is not a valid XLSX ({len(content)} bytes)")
    return content


def download_binary_tree(session: requests.Session) -> list[dict]:
    response = session.get(BINARY_TREE_URL, params={
        "userid": ROOT_MEMBER_ID,
        "type": "sponsor",
        "level": "100",
    }, headers={"Accept": "application/json"}, timeout=90)
    response.raise_for_status()
    payload = response.json()
    rows = payload.get("Data", []) if isinstance(payload, dict) else []
    if not isinstance(rows, list) or len(rows) < 100:
        raise RuntimeError("First Global Binary Tree returned an invalid response")
    if not any(str(row.get("USERID", "")) == ROOT_MEMBER_ID for row in rows):
        raise RuntimeError(f"Binary Tree root {ROOT_MEMBER_ID} is missing")
    return rows


def merge_binary_tree(members: dict, reports: list[dict], tree_rows: list[dict]) -> int:
    report_ids = {report["member_id"] for report in reports}
    tree_ids: set[str] = set()
    upline_counts: Counter = Counter()

    for row in tree_rows:
        member_id = str(row.get("USERID", "")).strip()
        if not member_id or member_id in tree_ids:
            raise RuntimeError(f"Invalid or duplicate Binary Tree member: {member_id!r}")
        tree_ids.add(member_id)

        upline_id = str(row.get("P_ID", "")).strip() or None
        sponsor_id = str(row.get("R_ID", "")).strip() or None
        if upline_id:
            upline_counts[upline_id] += 1

        connector = str(row.get("STATUS", "")).strip() == "0"
        member = members.get(member_id)
        if member is None:
            box_parts = str(row.get("BOX_TXT") or "").split("|")
            name = __import__("re").sub(r"\([LR]\)$", "", box_parts[0]).strip() if box_parts else member_id
            members[member_id] = {
                "id": member_id,
                "name": name or member_id,
                "join_date": box_parts[2] if len(box_parts) > 2 else None,
                "country": None,
                "lv": 0,
                "upline_id": upline_id,
                "sponsor_id": sponsor_id,
                "placement_connector": connector,
            }
        else:
            # The Binary Tree endpoint is authoritative for Placement and Sponsor.
            member["upline_id"] = upline_id
            member["sponsor_id"] = sponsor_id
            if connector:
                member["placement_connector"] = True
            else:
                member.pop("placement_connector", None)

    invalid_uplines = [member_id for member_id, count in upline_counts.items() if count > 2]
    if invalid_uplines:
        raise RuntimeError(f"Invalid Binary Tree placement under: {', '.join(invalid_uplines[:5])}")

    missing_report_members = report_ids - tree_ids
    if missing_report_members:
        raise RuntimeError(
            f"Business Report members missing from Binary Tree: {', '.join(sorted(missing_report_members)[:5])}"
        )
    return sum(1 for member in members.values() if member.get("placement_connector"))


def parse_report(content: bytes, month: str) -> tuple[dict, list[dict]]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_name = next(
        (name for name in ("Business Report", "รายงานธุรกิจ") if name in workbook.sheetnames),
        None,
    )
    if not sheet_name:
        raise RuntimeError("Business Report worksheet is missing")
    worksheet = workbook[sheet_name]
    is_first_global = str(worksheet.cell(1, 2).value or "").strip() == "รหัสสมาชิก"
    if is_first_global:
        sponsor_header = str(worksheet.cell(1, 14).value or "")
        upline_header = str(worksheet.cell(1, 15).value or "")
        relationships_valid = sponsor_header == "ผู้แนะนำ" and upline_header == "ผู้สนับสนุน"
    else:
        upline_header = str(worksheet.cell(1, 11).value or "")
        sponsor_header = str(worksheet.cell(1, 12).value or "")
        relationships_valid = "Upline" in upline_header and "Sponsor" in sponsor_header
    if not relationships_valid:
        raise RuntimeError(
            f"Unexpected relationship columns: upline={upline_header!r}, sponsor={sponsor_header!r}"
        )

    members: dict = {}
    reports = process_file(worksheet, month, members)
    ids = [report["member_id"] for report in reports]
    if not 100 <= len(reports) <= 10_000:
        raise RuntimeError(f"Unexpected report row count: {len(reports)}")
    if len(ids) != len(set(ids)):
        raise RuntimeError("Duplicate member ids in Business Report")
    if ROOT_MEMBER_ID not in members:
        raise RuntimeError(f"Root member {ROOT_MEMBER_ID} is missing")

    upline_counts = Counter(member.get("upline_id") for member in members.values() if member.get("upline_id"))
    invalid_uplines = [member_id for member_id, count in upline_counts.items() if count > 2]
    if invalid_uplines:
        raise RuntimeError(f"Invalid binary placement under upline: {', '.join(invalid_uplines[:5])}")

    root_sponsored = sum(1 for member in members.values() if member.get("sponsor_id") == ROOT_MEMBER_ID)
    root_placed = upline_counts.get(ROOT_MEMBER_ID, 0)
    if root_sponsored < 2 or root_placed > 2:
        raise RuntimeError(
            f"Relationship mapping check failed for {ROOT_MEMBER_ID}: "
            f"sponsored={root_sponsored}, placed={root_placed}"
        )
    return members, reports


def sync_report(sync_url: str, secret: str, payload: dict) -> dict:
    response = requests.post(
        sync_url,
        json=payload,
        headers={"Authorization": f"Bearer {secret}"},
        timeout=90,
    )
    if not response.ok:
        raise RuntimeError(f"Production sync failed ({response.status_code}): {response.text[:500]}")
    return response.json()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", help="Report month in YYYY-MM format")
    parser.add_argument("--include-previous", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--output-dir", type=Path, default=Path("downloaded_reports"))
    args = parser.parse_args()

    username = os.environ.get("FIRST_THAILAND_USER", "")
    password = os.environ.get("FIRST_THAILAND_PASSWORD", "")
    secret = os.environ.get("BUSINESS_REPORT_SYNC_SECRET", "")
    sync_url = os.environ.get(
        "BUSINESS_REPORT_SYNC_URL",
        "https://downline-analyzer.vercel.app/api/admin/business-report-sync",
    )
    if not username or not password:
        raise RuntimeError("FIRST_THAILAND_USER and FIRST_THAILAND_PASSWORD are required")
    if not args.dry_run and not secret:
        raise RuntimeError("BUSINESS_REPORT_SYNC_SECRET is required")
    if args.month and not __import__("re").fullmatch(r"\d{4}-(0[1-9]|1[0-2])", args.month):
        raise RuntimeError("--month must use YYYY-MM")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "POST"),
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    login(session, username, password)
    binary_tree = download_binary_tree(session)

    for month in selected_months(args.month, args.include_previous):
        content = download(session, month)
        checksum = hashlib.sha256(content).hexdigest()
        members, reports = parse_report(content, month)
        connector_count = merge_binary_tree(members, reports, binary_tree)
        output_path = args.output_dir / f"business_report_SPS_{month}.xlsx"
        output_path.write_bytes(content)
        payload = {
            "month": month,
            "checksum": checksum,
            "members": members,
            "reports": reports,
        }
        result = {
            "month": month,
            "rows": len(reports),
            "members": len(members),
            "placement_connectors": connector_count,
            "checksum": checksum,
        }
        if not args.dry_run:
            result["production"] = sync_report(sync_url, secret, payload)
        print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
